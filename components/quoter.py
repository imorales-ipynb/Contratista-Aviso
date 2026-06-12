import os
import streamlit as st
import pandas as pd
import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

EMAIL_EMPRESA    = "venta.ticket@casinoexpress.cl"
TELEFONOS        = "+569 42367538 / +569 69192409"
CONDICIONES_PAGO = ["Anticipado", "Tarjeta de Crédito", "Tarjeta de Débito",
                    "Crédito 30 días", "Crédito 60 días"]
OPERADORES       = ["Fabián Flores", "Héctor Astudillo", "Wilson Prado", "Werner Marti"]
LOGO_PATH        = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets", "logo.png")


# ── Helpers Excel ─────────────────────────────────────────────────────────────

def _celda(ws, fila, col, valor="", bold=False, size=10, color_font="000000",
           bg=None, alineacion="left", wrap=False, numero_fmt=None, borde=None):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font      = Font(name="Calibri", bold=bold, size=size, color=color_font)
    c.alignment = Alignment(horizontal=alineacion, vertical="center", wrap_text=wrap)
    if bg:        c.fill   = PatternFill(fill_type="solid", fgColor=bg)
    if numero_fmt: c.number_format = numero_fmt
    if borde:     c.border = borde
    return c

def _borde_fino():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Excel ─────────────────────────────────────────────────────────────────────

def _exportar_excel(df_cot, casino, fecha, vigencia, cliente, rut,
                    condicion_pago, numero="", operador="",
                    gerente_servicio="", jefe_servicio=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    AZUL    = "1F4E79"
    AZUL_CL = "D6E4F0"
    GRIS_H  = "595959"
    GRIS_F  = "F2F2F2"
    NARANJA = "C55A11"
    B       = _borde_fino()

    anchos = [5, 40, 22, 16, 18, 10, 16]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    # ── Fila 1: Logo + Título ─────────────────────────────────────────────────
    ws.row_dimensions[1].height = 55

    if os.path.exists(LOGO_PATH):
        from openpyxl.drawing.image import Image as XlImage
        logo        = XlImage(LOGO_PATH)
        logo.width  = 155
        logo.height = 52
        ws.add_image(logo, "A1")
        ws.merge_cells("A1:C1")

    titulo = f"COTIZACIÓN DE SERVICIOS   {numero}" if numero else "COTIZACIÓN DE SERVICIOS"
    ws.merge_cells("D1:G1")
    _celda(ws, 1, 4, titulo, bold=True, size=14, color_font=AZUL, alineacion="center")

    # ── Filas 2-7: Info cliente / documento ──────────────────────────────────
    contacto = f"{EMAIL_EMPRESA}   |   Tel: {TELEFONOS}"
    info_doc = [
        ("N° Cotizacion:", numero,          "Emision:",        fecha.strftime("%d/%m/%Y")),
        ("Cliente:",       cliente,         "Vigencia:",       vigencia.strftime("%d/%m/%Y")),
        ("RUT:",           rut,             "Cond. Pago:",     condicion_pago),
        ("Casino:",        casino,          "Operador:",       operador),
        ("Gte. Servicio:", gerente_servicio,"Jefe Servicio:",  jefe_servicio),
        ("Email / Tel:",   contacto,        "",                ""),
    ]
    for i, (l1, v1, l2, v2) in enumerate(info_doc, start=2):
        last = (i == 7)
        if last:
            ws.merge_cells(f"A{i}:G{i}")
            _celda(ws, i, 1, f"{l1}  {v1}", size=9, color_font="595959")
        else:
            ws.merge_cells(f"A{i}:B{i}"); ws.merge_cells(f"C{i}:D{i}")
            ws.merge_cells(f"E{i}:F{i}")
            _celda(ws, i, 1, f"{l1}  {v1}", bold=(i==2), size=10,
                   color_font=AZUL if i==2 else "000000")
            _celda(ws, i, 3, "")
            _celda(ws, i, 5, f"{l2}  {v2}", bold=(l2 != ""), size=10)
            _celda(ws, i, 7, "")
        ws.row_dimensions[i].height = 16

    ws.row_dimensions[8].height = 6

    # ── Tabla servicios ───────────────────────────────────────────────────────
    HEADERS    = ["N°", "Servicio", "Alias", "Cód. Servicio",
                  "Precio Unitario", "Cantidad", "Subtotal"]
    HEADER_ROW = 9
    for ci, h in enumerate(HEADERS, 1):
        _celda(ws, HEADER_ROW, ci, h, bold=True, size=10, color_font="FFFFFF",
               bg=AZUL, alineacion="center", borde=B)
    ws.row_dimensions[HEADER_ROW].height = 22

    fmt_pesos  = '$#,##0'
    data_start = HEADER_ROW + 1
    for idx, (_, row) in enumerate(df_cot.iterrows(), 1):
        r      = data_start + idx - 1
        bg_row = GRIS_F if idx % 2 == 0 else None
        vals   = [idx, row.get("NombreServicio",""), row.get("Alias",""),
                  row.get("Codigo Servicio",""), row.get("Precio",0),
                  row.get("Cantidad",0), row.get("Subtotal",0)]
        for ci, val in enumerate(vals, 1):
            alin = "right" if ci in (5,7) else ("center" if ci in (1,6) else "left")
            fmt  = fmt_pesos if ci in (5,7) else None
            _celda(ws, r, ci, val, size=10, bg=bg_row, alineacion=alin,
                   numero_fmt=fmt, borde=B)
        ws.row_dimensions[r].height = 16

    # ── Totales ───────────────────────────────────────────────────────────────
    neto          = df_cot["Subtotal"].sum()
    iva           = neto * 0.19
    total_con_iva = neto + iva
    tr            = data_start + len(df_cot)

    def _fila_total(fila, etiqueta, valor, dest=False):
        ws.merge_cells(f"A{fila}:F{fila}")
        _celda(ws, fila, 1, etiqueta, bold=dest, size=11,
               color_font=AZUL if dest else "000000",
               bg=AZUL_CL if dest else "EBF3FB",
               alineacion="right", borde=B)
        _celda(ws, fila, 7, valor, bold=dest, size=11,
               color_font=AZUL if dest else "000000",
               bg=AZUL_CL if dest else "EBF3FB",
               alineacion="right", numero_fmt=fmt_pesos, borde=B)
        ws.row_dimensions[fila].height = 18

    _fila_total(tr,   "Total Neto", neto)
    _fila_total(tr+1, "IVA (19%)",  iva)
    _fila_total(tr+2, "TOTAL",      total_con_iva, dest=True)

    # ── Pie informativo ───────────────────────────────────────────────────────
    h = tr + 5
    ws.merge_cells(f"A{h}:D{h}")
    _celda(ws, h, 1, "Datos de Transferencia", bold=True, size=10,
           color_font="FFFFFF", bg=GRIS_H, alineacion="left", borde=B)
    ws.row_dimensions[h].height = 18

    transferencia = ["Banco: Chile", "Cuenta Corriente: 167-01052-02",
                     "Rut: 78.793.360-2", "Casino Express S.A",
                     f"Mail: {EMAIL_EMPRESA}",
                     f"Tel: {TELEFONOS}"]
    for i, linea in enumerate(transferencia, 1):
        ws.merge_cells(f"A{h+i}:D{h+i}")
        _celda(ws, h+i, 1, linea, size=9, borde=B)
        ws.row_dimensions[h+i].height = 14

    texto_vigencia = (
        "Los tickets comprados tienen una vigencia de 100 días a contar de la fecha de "
        "emisión. Art. 41 Ley 19496. El prestador no efectuará devolución de dinero por "
        "no uso, salvo que el servicio no esté disponible en los días y horas en que se presta."
    )
    ws.merge_cells(f"E{h}:G{h+len(transferencia)}")
    c = ws.cell(row=h, column=5, value=texto_vigencia)
    c.font      = Font(name="Calibri", size=12, color=NARANJA)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = B

    h = h + len(transferencia) + 1
    ws.merge_cells(f"A{h}:D{h+4}")
    c = ws.cell(row=h, column=1,
                value="Los pagos realizados posterior a las 14:00 horas serán considerados "
                      "como pagos del día siguiente, incluyendo pagos informados al correo "
                      "posterior al horario indicado.")
    c.font      = Font(name="Calibri", size=9, italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = B

    ws.merge_cells(f"E{h}:G{h}")
    _celda(ws, h, 5,
           "Estimado cliente, para recibir nuestros servicios debe realizar el pago con anticipación:",
           bold=True, size=9, color_font="FFFFFF", bg=GRIS_H,
           alineacion="left", wrap=True, borde=B)
    ws.row_dimensions[h].height = 28

    for j, inst in enumerate([
        "Pago mediante transferencia electrónica 24 horas hábiles de anticipación.",
        "Pago mediante GetNet 48 horas hábiles de anticipación.",
        "Se solicita programar sus pagos para evitar suspensión de los servicios.",
    ], 1):
        ws.merge_cells(f"E{h+j}:G{h+j}")
        _celda(ws, h+j, 5, inst, size=9, color_font=NARANJA, borde=B)
        ws.row_dimensions[h+j].height = 14

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ── PDF ───────────────────────────────────────────────────────────────────────

def _exportar_pdf(df_cot, casino, fecha, vigencia, cliente, rut,
                  condicion_pago, numero="", operador="",
                  gerente_servicio="", jefe_servicio=""):
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_margins(15, 15, 15)
    W = 180

    AZUL    = (31, 78, 121)
    AZUL_CL = (214, 228, 240)
    NARANJA = (197, 90, 17)
    GRIS_H  = (89, 89, 89)
    GRIS_F  = (242, 242, 242)

    # ── Logo + Título ─────────────────────────────────────────────────────────
    if os.path.exists(LOGO_PATH):
        pdf.image(LOGO_PATH, x=15, y=8, w=52)
        pdf.set_y(40)
    else:
        pdf.set_y(15)

    titulo = f"COTIZACION DE SERVICIOS   {numero}" if numero else "COTIZACION DE SERVICIOS"
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_text_color(*AZUL)
    pdf.cell(W, 10, titulo, border=0, ln=1, align="C")
    pdf.ln(5)

    # ── Info cliente / documento ──────────────────────────────────────────────
    info_rows = [
        ("N Cotizacion:", numero or "-",        "Emision:",       fecha.strftime("%d/%m/%Y")),
        ("Cliente:",      cliente or "-",       "Vigencia:",      vigencia.strftime("%d/%m/%Y")),
        ("RUT:",          rut or "-",           "Cond. Pago:",    condicion_pago),
        ("Casino:",       casino,               "Operador:",      operador or "-"),
        ("Gte. Servicio:",gerente_servicio or "-","Jefe Servicio:", jefe_servicio or "-"),
    ]
    L, V = 26, 64
    H_ROW = 7
    for l1, v1, l2, v2 in info_rows:
        pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(0, 0, 0)
        pdf.cell(L, H_ROW, l1, ln=0)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(V, H_ROW, v1, ln=0)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(L, H_ROW, l2, ln=0)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(V, H_ROW, v2, ln=1)
        pdf.ln(1)

    # Fila contacto completa
    pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(0, 0, 0)
    pdf.cell(L, H_ROW, "Email / Tel:", ln=0)
    pdf.set_font("Helvetica", "", 8); pdf.set_text_color(89, 89, 89)
    pdf.cell(W - L, H_ROW, f"{EMAIL_EMPRESA}   |   Tel: {TELEFONOS}", ln=1)
    pdf.ln(5)

    # ── Encabezado tabla ──────────────────────────────────────────────────────
    HEADERS = ["N", "Servicio", "Codigo", "Precio Unit.", "Cant.", "Subtotal"]
    COL_W   = [8, 80, 22, 28, 14, 28]
    ALIGNS  = ["C", "L", "C", "R", "C", "R"]

    pdf.set_fill_color(*AZUL); pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    for h, w, a in zip(HEADERS, COL_W, ALIGNS):
        pdf.cell(w, 8, h, border=1, ln=0, align=a, fill=True)
    pdf.ln()

    # ── Filas de servicios ────────────────────────────────────────────────────
    neto = 0.0
    for idx, (_, row) in enumerate(df_cot.iterrows(), 1):
        subtotal = float(row.get("Cantidad", 0)) * float(row.get("Precio", 0))
        neto += subtotal
        fill = (idx % 2 == 0)
        pdf.set_fill_color(*(GRIS_F if fill else (255, 255, 255)))
        pdf.set_text_color(0, 0, 0); pdf.set_font("Helvetica", "", 9)
        nombre = str(row.get("NombreServicio", ""))
        if len(nombre) > 55: nombre = nombre[:52] + "..."
        vals = [str(idx), nombre,
                str(row.get("Codigo Servicio","")),
                f"${float(row.get('Precio',0)):,.0f}",
                str(int(row.get("Cantidad",0))), f"${subtotal:,.0f}"]
        for val, w, a in zip(vals, COL_W, ALIGNS):
            pdf.cell(w, 7, val, border=1, ln=0, align=a, fill=True)
        pdf.ln()

    # ── Totales ───────────────────────────────────────────────────────────────
    iva   = neto * 0.19
    total = neto + iva
    TW    = sum(COL_W[:-1])

    def _total(label, valor, dest=False):
        if dest:
            pdf.set_fill_color(*AZUL_CL); pdf.set_text_color(*AZUL)
            pdf.set_font("Helvetica", "B", 10)
        else:
            pdf.set_fill_color(235, 243, 251); pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "", 9)
        pdf.cell(TW, 7, label, border=1, ln=0, align="R", fill=True)
        pdf.cell(COL_W[-1], 7, f"${valor:,.0f}", border=1, ln=1, align="R", fill=True)

    _total("Total Neto", neto)
    _total("IVA (19%)", iva)
    _total("TOTAL", total, dest=True)
    pdf.ln(5)

    # ── Pie ───────────────────────────────────────────────────────────────────
    y0 = pdf.get_y(); xL = pdf.l_margin; xR = xL + 93; cw = 87

    pdf.set_fill_color(*GRIS_H); pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_xy(xL, y0); pdf.cell(cw, 7, "Datos de Transferencia", border=1, ln=0, fill=True)
    pdf.set_xy(xR, y0); pdf.cell(cw, 7, "Instrucciones de Pago",  border=1, ln=1, fill=True)

    y1 = pdf.get_y()
    transferencia = ["Banco: Chile", "Cuenta Corriente: 167-01052-02",
                     "Rut: 78.793.360-2", "Casino Express S.A",
                     f"Mail: {EMAIL_EMPRESA}",
                     f"Tel: {TELEFONOS}"]
    pdf.set_text_color(0, 0, 0); pdf.set_font("Helvetica", "", 8)
    for i, linea in enumerate(transferencia):
        pdf.set_xy(xL, y1 + i * 5.5); pdf.cell(cw, 5.5, linea, border=1)

    pdf.set_text_color(*NARANJA); pdf.set_font("Helvetica", "", 8)
    instrucciones = ["Para recibir servicios, pague con anticipacion:",
                     "- Transferencia: 24 hrs habiles.",
                     "- Getnet: 48 hrs habiles.",
                     "- Programe pagos para evitar suspension.", ""]
    for i, inst in enumerate(instrucciones):
        pdf.set_xy(xR, y1 + i * 5.5); pdf.cell(cw, 5.5, inst, border=1)

    # ── Aviso vigencia al final de la página ──────────────────────────────────
    pdf.set_auto_page_break(False)
    pdf.set_y(-28)
    pdf.set_fill_color(*NARANJA)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 10)
    pdf.multi_cell(W, 7,
        "IMPORTANTE: Los tickets comprados tienen una vigencia de 100 dias a contar de la fecha "
        "de emision. Art. 41 Ley 19496. El prestador no efectuara devolucion de dinero por no "
        "uso, salvo que el servicio no este disponible en los dias y horas en que se presta.",
        border=0, align="L", fill=True)

    return bytes(pdf.output())


# ── UI Cotizador ──────────────────────────────────────────────────────────────

def _buscar_jerarquia(df_jerarquia, casino_sel):
    """Devuelve (gerente, jefe) para el casino dado."""
    if df_jerarquia is None or df_jerarquia.empty:
        return "", ""
    fila = df_jerarquia[
        df_jerarquia["Casino"].str.strip().str.lower() == casino_sel.strip().lower()
    ]
    if fila.empty:
        return "", ""
    row = fila.iloc[0]
    return str(row.get("GOP", "") or ""), str(row.get("JOP", "") or "")


def render_cotizador(df_precios, df_clientes=None, df_jerarquia=None):
    st.title("Cotizador de Servicios")

    if "cot_items" not in st.session_state:
        st.session_state.cot_items = []
    if "cot_casino_prev" not in st.session_state:
        st.session_state.cot_casino_prev = None

    if df_precios.empty:
        st.warning("No hay datos de precios cargados.")
        if st.button("🔄 Cargar Precios desde FasesUAT"):
            from database import recargar_precios
            with st.spinner("Descargando precios..."):
                recargar_precios()
            st.rerun()
        return

    hoy            = datetime.date.today()
    vigencia_fecha = hoy + datetime.timedelta(days=7)

    # ── Casino ────────────────────────────────────────────────────────────────
    col_cas, col_reload = st.columns([6, 1])
    with col_cas:
        casinos_disp = sorted(df_precios["Nombre Casino"].dropna().unique().tolist())
        casino_sel   = st.selectbox("Casino:", options=casinos_disp, key="cot_casino")
    with col_reload:
        st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar Precios", key="cot_reload"):
            from database import recargar_precios, recargar_clientes
            with st.spinner("Actualizando..."): recargar_precios(); recargar_clientes()
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    if st.session_state.cot_casino_prev != casino_sel:
        st.session_state.cot_items = []
        st.session_state.cot_casino_prev = casino_sel
        for k in ("cot_cliente_sel", "cot_rut_sel", "cot_nombre_manual",
                  "cot_rut_manual", "cot_creada", "cot_estado_hash"):
            st.session_state.pop(k, None)

    # ── Operador + Jerarquía ──────────────────────────────────────────────────
    gerente, jefe = _buscar_jerarquia(df_jerarquia, casino_sel)

    col_op, col_gte, col_jefe = st.columns(3)
    with col_op:
        operador = st.selectbox("Operador:", options=OPERADORES, key="cot_operador")
    with col_gte:
        st.text_input("Gerente de Servicio:", value=gerente, disabled=True, key=f"cot_gerente_{casino_sel}")
    with col_jefe:
        st.text_input("Jefe de Servicio:", value=jefe, disabled=True, key=f"cot_jefe_{casino_sel}")

    # ── Datos del Cliente ─────────────────────────────────────────────────────
    st.markdown("#### Datos del Cliente")
    manual = st.checkbox("Ingresar cliente manualmente (no está en el listado)",
                         key="cot_manual", value=False)

    if manual:
        col_c1, col_c2, col_c3 = st.columns(3)
        with col_c1:
            nombre_manual = st.text_input("Nombre del Cliente:", key="cot_nombre_manual",
                                          placeholder="Razón social...")
        with col_c2:
            rut_manual = st.text_input("RUT:", key="cot_rut_manual",
                                       placeholder="XX.XXX.XXX-X")
        with col_c3:
            condicion = st.selectbox("Condiciones de Pago:", options=CONDICIONES_PAGO,
                                     key="cot_condicion")
        cliente_val = nombre_manual.strip()
        rut_auto    = rut_manual.strip()
    else:
        df_uniq = pd.DataFrame()
        if df_clientes is not None and not df_clientes.empty:
            df_uniq = (df_clientes[["RazonSocial", "RutContratista"]]
                       .drop_duplicates().sort_values("RazonSocial").reset_index(drop=True))

        PH_N = "— Buscar por nombre —"
        PH_R = "— Buscar por RUT —"
        op_n = [PH_N] + (df_uniq["RazonSocial"].tolist() if not df_uniq.empty else [])
        op_r = [PH_R] + (sorted(df_uniq["RutContratista"].astype(str).unique().tolist())
                          if not df_uniq.empty else [])

        if st.session_state.get("cot_cliente_sel", PH_N) not in op_n:
            st.session_state["cot_cliente_sel"] = PH_N
        if st.session_state.get("cot_rut_sel", PH_R) not in op_r:
            st.session_state["cot_rut_sel"] = PH_R

        def _on_nombre():
            n = st.session_state.get("cot_cliente_sel", PH_N)
            if n != PH_N and not df_uniq.empty:
                m = df_uniq[df_uniq["RazonSocial"] == n]
                st.session_state["cot_rut_sel"] = str(m["RutContratista"].iloc[0]) if not m.empty else PH_R
            else:
                st.session_state["cot_rut_sel"] = PH_R

        def _on_rut():
            r = st.session_state.get("cot_rut_sel", PH_R)
            if r != PH_R and not df_uniq.empty:
                m = df_uniq[df_uniq["RutContratista"].astype(str) == r]
                st.session_state["cot_cliente_sel"] = m["RazonSocial"].iloc[0] if not m.empty else PH_N
            else:
                st.session_state["cot_cliente_sel"] = PH_N

        col_c1, col_c2, col_c3 = st.columns(3)
        with col_c1:
            nombre_sel = st.selectbox("Nombre del Cliente:", options=op_n,
                                      key="cot_cliente_sel", on_change=_on_nombre)
        with col_c2:
            rut_sel = st.selectbox("RUT:", options=op_r,
                                   key="cot_rut_sel", on_change=_on_rut)
        with col_c3:
            condicion = st.selectbox("Condiciones de Pago:", options=CONDICIONES_PAGO,
                                     key="cot_condicion")

        cliente_val = nombre_sel if nombre_sel != PH_N else ""
        rut_auto    = rut_sel    if rut_sel    != PH_R else ""

    col_d1, col_d2, col_d3 = st.columns(3)
    with col_d1:
        st.text_input("Email de contacto:", value=EMAIL_EMPRESA, disabled=True, key="cot_email")
    with col_d2:
        st.text_input("Fecha de Emisión:", value=hoy.strftime("%d/%m/%Y"),
                      disabled=True, key="cot_emision")
    with col_d3:
        st.text_input("Vigencia:", value=vigencia_fecha.strftime("%d/%m/%Y"),
                      disabled=True, key="cot_vigencia")

    # ── Buscador de servicios ─────────────────────────────────────────────────
    st.markdown("---")
    COLS_ITEM        = ["NombreServicio", "TipoServicio", "Alias", "Codigo Servicio", "Precio"]
    servicios_casino = (df_precios[df_precios["Nombre Casino"] == casino_sel]
                        [[c for c in COLS_ITEM if c in df_precios.columns]]
                        .drop_duplicates())
    ya_agregados     = {i["NombreServicio"] for i in st.session_state.cot_items}
    servicios_disp   = servicios_casino[~servicios_casino["NombreServicio"].isin(ya_agregados)]

    col_search, col_add = st.columns([5, 1])
    with col_search:
        if servicios_disp.empty:
            st.info("Todos los servicios de este casino ya están agregados.")
            servicio_sel = None
        else:
            servicio_sel = st.selectbox("Buscar servicio:",
                                        options=sorted(servicios_disp["NombreServicio"].unique().tolist()),
                                        key="cot_search")
    with col_add:
        st.markdown("<div style='margin-top:28px'>", unsafe_allow_html=True)
        agregar = st.button("➕ Agregar", use_container_width=True, key="cot_agregar")
        st.markdown("</div>", unsafe_allow_html=True)

    if agregar and servicio_sel:
        fila = servicios_disp[servicios_disp["NombreServicio"] == servicio_sel].iloc[0]
        st.session_state.cot_items.append({
            "NombreServicio":  fila.get("NombreServicio", ""),
            "TipoServicio":    fila.get("TipoServicio", ""),
            "Alias":           fila.get("Alias", ""),
            "Codigo Servicio": fila.get("Codigo Servicio", ""),
            "Precio":          fila.get("Precio", 0),
            "Cantidad":        1,
        })
        st.session_state.pop("cot_creada", None)
        st.rerun()

    # ── Vista previa ──────────────────────────────────────────────────────────
    st.markdown("#### Vista previa de cotización")

    if not st.session_state.cot_items:
        st.info("Selecciona un servicio y presiona **➕ Agregar** para comenzar.")
        return

    df_preview         = pd.DataFrame(st.session_state.cot_items)
    df_preview["Eliminar"] = False

    df_editado = st.data_editor(
        df_preview,
        column_config={
            "NombreServicio":  st.column_config.TextColumn("Servicio",      disabled=True, width="large"),
            "Alias":           st.column_config.TextColumn("Alias",         disabled=True),
            "Codigo Servicio": st.column_config.TextColumn("Código",        disabled=True),
            "Precio":          st.column_config.NumberColumn("Precio Unit.", disabled=True, format="$%d"),
            "Cantidad":        st.column_config.NumberColumn("Cantidad",    min_value=1, step=1, format="%d"),
            "Eliminar":        st.column_config.CheckboxColumn("🗑️"),
        },
        column_order=["NombreServicio", "Alias", "Codigo Servicio", "Precio", "Cantidad", "Eliminar"],
        hide_index=True,
        use_container_width=True,
        key=f"cot_tabla_{casino_sel}_{len(st.session_state.cot_items)}",
    )

    for i, row in df_editado.iterrows():
        if i < len(st.session_state.cot_items):
            qty = row.get("Cantidad", 1)
            st.session_state.cot_items[i]["Cantidad"] = max(1, int(qty) if pd.notna(qty) else 1)

    col_del, col_clear, _ = st.columns([1, 1, 5])
    with col_del:
        if st.button("Eliminar marcados", key="cot_del"):
            st.session_state.cot_items = [
                item for i, item in enumerate(st.session_state.cot_items)
                if i >= len(df_editado) or not df_editado.iloc[i].get("Eliminar", False)
            ]
            st.session_state.pop("cot_creada", None)
            st.rerun()
    with col_clear:
        if st.button("Limpiar todo", key="cot_clear"):
            st.session_state.cot_items = []
            st.session_state.pop("cot_creada", None)
            st.rerun()

    # ── Totales ───────────────────────────────────────────────────────────────
    df_cot = df_editado.drop(columns=["Eliminar"]).copy()
    df_cot["Subtotal"] = df_cot["Cantidad"] * df_cot["Precio"]

    neto_ui  = df_cot["Subtotal"].sum()
    iva_ui   = neto_ui * 0.19
    total_ui = neto_ui + iva_ui

    col_k1, col_k2, col_k3, col_k4, col_k5 = st.columns(5)
    with col_k1: st.metric("Servicios",       len(df_cot))
    with col_k2: st.metric("Tickets totales", f"{int(df_cot['Cantidad'].sum()):,}")
    with col_k3: st.metric("Total Neto",      f"${neto_ui:,.0f}")
    with col_k4: st.metric("IVA (19%)",       f"${iva_ui:,.0f}")
    with col_k5: st.metric("Total c/ IVA",    f"${total_ui:,.0f}")

    # ── Detectar cambios para invalidar cotización ya creada ──────────────────
    estado = f"{casino_sel}|{cliente_val}|{rut_auto}|{condicion}|" \
             + str([(i["NombreServicio"], i["Cantidad"]) for i in st.session_state.cot_items])
    if st.session_state.get("cot_estado_hash") != estado:
        st.session_state["cot_estado_hash"] = estado
        st.session_state.pop("cot_creada", None)

    st.markdown("---")

    # ── Flujo: Crear → Descargar ──────────────────────────────────────────────
    if not st.session_state.get("cot_creada", False):
        st.info("Revisa los datos y presiona **Crear Cotización** para registrarla y habilitar la descarga.")
        if st.button("✅ Crear Cotización", type="primary", use_container_width=True, key="cot_crear"):
            from utils.historial import guardar_cotizacion, _siguiente_numero, formato_numero
            n      = _siguiente_numero()
            numero = formato_numero(n, hoy.year)

            export_args = dict(df_cot=df_cot, casino=casino_sel, fecha=hoy,
                               vigencia=vigencia_fecha, cliente=cliente_val,
                               rut=rut_auto, condicion_pago=condicion, numero=numero,
                               operador=operador, gerente_servicio=gerente,
                               jefe_servicio=jefe)

            with st.spinner("Generando documentos..."):
                excel_data = _exportar_excel(**export_args)
                pdf_data   = _exportar_pdf(**export_args)
                guardar_cotizacion(**export_args, excel_bytes=excel_data, pdf_bytes=pdf_data)

            st.session_state["cot_creada"]       = True
            st.session_state["cot_numero"]        = numero
            st.session_state["cot_excel_dl"]      = excel_data
            st.session_state["cot_pdf_dl"]        = pdf_data
            st.session_state["cot_nombre_base"]   = (
                f"Cotizacion_{numero}_{casino_sel.replace(' ','_')}")
            st.rerun()
    else:
        numero     = st.session_state.get("cot_numero", "")
        nombre_base = st.session_state.get("cot_nombre_base", f"Cotizacion_{numero}")

        st.success(f"Cotización **{numero}** creada y guardada en el historial.")

        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button("📥 Descargar Excel",
                               data=st.session_state["cot_excel_dl"],
                               file_name=f"{nombre_base}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with col_dl2:
            st.download_button("📄 Descargar PDF",
                               data=st.session_state["cot_pdf_dl"],
                               file_name=f"{nombre_base}.pdf",
                               mime="application/pdf",
                               use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🆕 Nueva Cotización", use_container_width=True, key="cot_nueva"):
            st.session_state.cot_items = []
            for k in ["cot_creada", "cot_numero", "cot_excel_dl", "cot_pdf_dl",
                      "cot_nombre_base", "cot_estado_hash", "cot_casino_prev"]:
                st.session_state.pop(k, None)
            st.rerun()
